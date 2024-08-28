from flask import Flask, render_template, request, redirect, url_for
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

app = Flask(__name__)

@app.route('/')
def reservation_form():
    return render_template('reservations.html')

@app.route('/submit', methods=['POST'])
def submit_form():
    if request.method == 'POST':

        # Récupération des données du formulaire
        nom = request.form['nom']
        telephone = request.form['phone']
        personnes = request.form['num_people']
        menu_poulet = request.form['poulet_royal']
        menu_tajine = request.form['tajine']
        menu_harira = request.form['harira']
        date = request.form['date']
        midi_soir = request.form['midi-soir']

        # Chemin du fichier Excel
        chemin_fichier_excel = r'C:\Users\33766\Desktop\reservations.xlsx'

        # Vérification et création du fichier Excel s'il n'existe pas
        if not os.path.exists(chemin_fichier_excel):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Réservations"
            # Ajouter des en-têtes au fichier Excel
            sheet.append(["Nom", "Téléphone", "Nombre de personnes", "Poulet Royal", "Tajine", "Harira", "Date", "Midi/Soir"])
            workbook.save(chemin_fichier_excel)
        else:
            workbook = openpyxl.load_workbook(chemin_fichier_excel)
            sheet = workbook.active

        # Ajout des données au tableau Excel
        row_data = [nom, telephone, personnes, menu_poulet, menu_tajine, menu_harira, date, midi_soir]
        sheet.append(row_data)
        workbook.save(chemin_fichier_excel)

        # Définir les styles que vous voulez appliquer
        fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
        font = Font(name='Calibri', size=12, bold=False)
        alignment = Alignment(horizontal='center', vertical='center')

        # Appliquer les styles à toutes les lignes à partir de la deuxième ligne
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fill
                cell.font = font
                cell.alignment = alignment

        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 20
        
        workbook.save(chemin_fichier_excel)
        return redirect(url_for('reservation_form'))

if __name__ == '__main__':
    app.run(debug=True)